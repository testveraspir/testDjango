import os
from django.shortcuts import render
from openpyxl import load_workbook


def read_movies_form_excel(file_path):
    movies = []

    workbook = load_workbook(filename=file_path)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        title, year, director, genre = row
        movie = {
            'title': title,
            'year': year,
            'director': director,
            'genre': genre
        }
        movies.append(movie)  # список, состоящий из словарей
    return movies


def movies_view(request):
    file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'files/movies.xlsx')
    movies = read_movies_form_excel(file_path)
    return render(request, 'index.html', {'movies': movies})
